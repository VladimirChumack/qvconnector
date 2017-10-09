# Qlikview connector using file selector example
# Minimalistic Qlikview connector example
# using python 2.7
# Author Vladimir Chumack
# https://www.linkedin.com/in/vladimirchumack/

import sys,os,logging, win32pipe, win32file ,wx, arrow, datetime, struct
from openpyxl import load_workbook
from qvconnector.QlikConnector import QlikConnector;

app = wx.App(None)

def get_path(wildcard):
    "File Selector"
    style = wx.FD_OPEN | wx.FD_FILE_MUST_EXIST
    dialog = wx.FileDialog(None, 'Open', wildcard=wildcard, style=style)
    if dialog.ShowModal() == wx.ID_OK:
        path = dialog.GetPath()
    else:
        path = 'Unknown'
    dialog.Destroy()
    return path

def getFields():
    "This procedure populates list of fields"
    logging.info(connector.execution_source);
    wb = load_workbook(connector.execution_source)
    ws = wb.active
    field_names =[];
    #reading first row in excel file
    for row in ws.iter_rows(min_row=1,max_row=1):
      for cell in row:
          field_names.append(cell.value.encode('utf-8'))        
    logging.info(field_names);
    connector.field_names=field_names;

def getScriptParameters():
    "This procedure populates file name and table name"
    connector.file_name  = get_path('*.xlsx').encode('utf-8');
    connector.table_name = 'table';

def getData():
    "This procedure loads the data"
    wb = load_workbook(connector.execution_source)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
      for cell in row:
          connector.writeStringField(str(cell.value));        
 
# Logging: Please change this path    
logging.basicConfig(filename=R'C:\Python27\connector\connector.log',level=logging.INFO);

# Creating connector
connector=QlikConnector();

# Running connector
connector.run(getFields,getScriptParameters,getData);    
    
#app = wx.App(None)
